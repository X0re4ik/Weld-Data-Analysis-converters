from datetime import date
from pathlib import Path

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.chart import (
    Reference,
    Series,
    ScatterChart
)
from openpyxl import Workbook

from notification_system.converters._converter import _Converter

class MeasurementsToExcelConverter(_Converter):
    def __init__(self, mac_address: str, date: date) -> None:
        _Converter.__init__(self, mac_address, date, "xlsx")
        
        self._work_book = Workbook()
        self._work_page: Worksheet = self._work_book.active
    
    def load_from_DB(self) -> bool:
        return _Converter.load_from_DB(self)
    
    def creat_file(self) -> None:
        _Converter.creat_file(self)
        
        
        START_ROW_FOR_WRITE_MEASUREMENTS = 13
        END_ROW_FOR_WRITE_MEASUREMENTS = START_ROW_FOR_WRITE_MEASUREMENTS + len(self.data_store.measurements) - 1
        
        START_COLUMN_FOR_WRITE_MEASUREMENTS = 2
        END_COLUMN_FOR_WRITE_MEASUREMENTS = START_COLUMN_FOR_WRITE_MEASUREMENTS + 3
        
        
        # set title for measure
        titles_and_ms = [
            ('время', ""),
            ("ток", "А"),
            ("расход газа", "л/мин"),
            ("расход проволки", "кг/час")
        ]
        
        for i, title_and_ms in enumerate(titles_and_ms):
            title = title_and_ms[0] + ' ' + title_and_ms[1]
            self._work_page.cell(column=START_COLUMN_FOR_WRITE_MEASUREMENTS+i, row=START_ROW_FOR_WRITE_MEASUREMENTS-1).value = title
        
        for i, measurement in enumerate(self.data_store.measurements):
            self._work_page.cell(column=START_COLUMN_FOR_WRITE_MEASUREMENTS, row=START_ROW_FOR_WRITE_MEASUREMENTS+i).value = measurement.utc_time
            self._work_page.cell(column=START_COLUMN_FOR_WRITE_MEASUREMENTS+1, row=START_ROW_FOR_WRITE_MEASUREMENTS+i).value = measurement.amperage
            self._work_page.cell(column=START_COLUMN_FOR_WRITE_MEASUREMENTS+2, row=START_ROW_FOR_WRITE_MEASUREMENTS+i).value = measurement.gas_consumption
            self._work_page.cell(column=END_COLUMN_FOR_WRITE_MEASUREMENTS, row=START_ROW_FOR_WRITE_MEASUREMENTS+i).value = measurement.wire_consumption
        
    
        
        from typing import List, Tuple
        def write_general_parameters(*,
                list_: List[Tuple[str, str]],
                name: str,
                command: str,
                cell_column, 
                cell_row, indent = 0):
            
            for i, (title, ms) in enumerate(list_):
                title = f"{name} {title}, {ms}"
                self._work_page.cell(column=cell_column, row=cell_row+i).value = title
                start_coordinate = self._work_page.cell(column=START_COLUMN_FOR_WRITE_MEASUREMENTS+1+indent+i, row=START_ROW_FOR_WRITE_MEASUREMENTS).coordinate
                end_coordinate = self._work_page.cell(column=START_COLUMN_FOR_WRITE_MEASUREMENTS+1+indent+i, row=END_ROW_FOR_WRITE_MEASUREMENTS).coordinate
                self._work_page.cell(column=cell_column+1, row=cell_row+i).value = f'={command}({start_coordinate}:{end_coordinate})'
        
        
        write_general_parameters(
            list_=titles_and_ms[1:],
            name='Средний',
            command='AVERAGE',
            cell_column=2,
            cell_row=3)
        
        write_general_parameters(
            list_=titles_and_ms[1:],
            name='Максимальный',
            command='MAX',
            cell_column=2,
            cell_row=6)
        
        write_general_parameters(
            list_=titles_and_ms[2:],
            name='Общий',
            command='SUM',
            cell_column=2,
            cell_row=9, indent=1)
        
        
        
        # # calculate AVERAGE values
        # for i, (title, ms) in enumerate(titles_and_ms):
        #     title = "Средний " + title + ', ' + ms
        #     self._work_page.cell(column=2, row=3+i).value = title
        #     start_coordinate = self._work_page.cell(column=START_COLUMN_FOR_WRITE_MEASUREMENTS+1+i, row=START_ROW_FOR_WRITE_MEASUREMENTS).coordinate
        #     end_coordinate = self._work_page.cell(column=START_COLUMN_FOR_WRITE_MEASUREMENTS+1+i, row=END_ROW_FOR_WRITE_MEASUREMENTS).coordinate
        #     self._work_page.cell(column=3, row=3+i).value = f'=AVERAGE({start_coordinate}:{end_coordinate})'
        
        # self.data_store.report.
        
        wpp = self.data_store.welding_process_parameters
        general_parameters = [
            ("Тип металла", wpp.metal_type),
            ("Тип газа", wpp.gas_type),
            ("Плотность металла, кг/м^3", wpp.metal_density),
            ("Диаметр проволки, мм", wpp.wire_diameter),
            
        ]
        for i, (title, value) in enumerate(general_parameters):
            self._work_page.cell(column=4, row=3+i).value = title
            self._work_page.cell(column=5, row=3+i).value = value
            
            
        
        x_time = Reference(
            self._work_page,
            min_col=START_COLUMN_FOR_WRITE_MEASUREMENTS, 
            min_row=START_ROW_FOR_WRITE_MEASUREMENTS, max_row=END_ROW_FOR_WRITE_MEASUREMENTS)
        
        for i, (title, ms) in enumerate(titles_and_ms[1:]):
            y = Reference(
                self._work_page, 
                min_col=START_COLUMN_FOR_WRITE_MEASUREMENTS+1+i, 
                min_row=START_ROW_FOR_WRITE_MEASUREMENTS, max_row=END_ROW_FOR_WRITE_MEASUREMENTS)
            
            s = Series(y, xvalues=x_time)
            chart = ScatterChart()
            chart.title = title
            chart.x_axis.title = ms
            chart.y_axis.title = 'Время'
            chart.legend = None
            chart.append(s)
            self._work_page.add_chart(chart, f"G{(i+1)*20}")
        
        
    def save(self, path: Path):
        self._work_book.save(self.name_file)
        