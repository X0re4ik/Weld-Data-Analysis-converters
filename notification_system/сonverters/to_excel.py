from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.chart import (
    Reference,
    Series,
    ScatterChart
)
from openpyxl import Workbook
from datetime import datetime, date, timedelta
from sqlalchemy import and_, func
from pathlib import Path

from notification_system.сonverters._converter import _Converter
from notification_system.models import (
    session,
    Sensor, DailyReport,
    Measurement, WeldMetal,
    WeldingWireDiameter, Worker
)

class MeasurementsToExcelConverter(_Converter):
    def __init__(self, mac_address: str, date: date) -> None:
        _Converter.__init__(self, mac_address, date, "xlsx")
        
        self._work_book = Workbook()
        self._work_page: Worksheet = self._work_book.active        
        
        self._option_for_measurement = {
            "titles": ["Время", "Сила тока", "Расход газа", "Расход проволки"],
            "ms": ["", "А", "л/мин", "кг/час"]
        }
        self._option_for_daily_report = {
            "average_amperage": {
                "ru": "Средняя сила тока",
                "value": "",
                "ms": "А"
            },
            "average_gas_consumption": {
                "ru": "Средний расход газа",
                "value": "",
                "ms": "л/мин"
            },
            "average_wire_consumption": {
                "ru": "Средний расход сварочной проволки",
                "value": "",
                "ms": "кг/час"
            },
            "date": {
                "ru": "День",
                "value": "",
                "ms": ""
            },
            "expended_gas": {
                "ru": "Израсходовано газа",
                "value": "",
                "ms": "л"
            },
            "expended_wire": {
                "ru": "Израсходовано проволки",
                "value": "",
                "ms": "кг"
            },
            "idle_time_in_seconds": {
                "ru": "Время простоя",
                "value": "",
                "ms": "с"
            },
            "max_amperage": {
                "ru": "Максимальная сила тока",
                "value": "",
                "ms": "А"
            },
            "max_gas_consumption": {
                "ru": "Максимальный расход газа",
                "value": "",
                "ms": "л/мин"
            },
            "max_wire_consumption": {
                "ru": "Максимальный расход проволки",
                "value": "",
                "ms": "кг/мин"
            },
            "running_time_in_seconds": {
                "ru": "Время работы",
                "value": "",
                "ms": "с"
            },
            "weld_metal": {
                "ru": "Материал проволки",
                "value": "",
                "ms": ""
            },
            "welding_wire_diameter": {
                "ru": "Диаметр проволки",
                "value": "",
                "ms": "мм"
            },
            "worker": {
                "ru": "Рабочий",
                "value": "",
                "ms": ""
            }
        }

        
    
    def load_from_DB(self) -> bool:
        return _Converter.load_from_DB(self)
    
    def creat_file(self) -> None:
        _Converter.creat_file(self)
        
        
        START_ROW_FOR_WRITE_MEASUREMENTS = 13
        END_ROW_FOR_WRITE_MEASUREMENTS = START_ROW_FOR_WRITE_MEASUREMENTS + len(self.data_store.measurements) - 1
        
        START_COLUMN_FOR_WRITE_MEASUREMENTS = 2
        END_COLUMN_FOR_WRITE_MEASUREMENTS = START_COLUMN_FOR_WRITE_MEASUREMENTS + 3
        
        
        # set title for measure
        titles_and_ms = [
            ('время', ""),
            ("ток", "А"),
            ("расход газа", "л/мин"),
            ("расход проволки", "кг/час")
        ]
        
        for i, title_and_ms in enumerate(titles_and_ms):
            title = title_and_ms[0] + ' ' + title_and_ms[1]
            self._work_page.cell(column=START_COLUMN_FOR_WRITE_MEASUREMENTS+i, row=START_ROW_FOR_WRITE_MEASUREMENTS-1).value = title
        
        for i, measurement in enumerate(self.data_store.measurements):
            self._work_page.cell(column=START_COLUMN_FOR_WRITE_MEASUREMENTS, row=START_ROW_FOR_WRITE_MEASUREMENTS+i).value = measurement.utc_time
            self._work_page.cell(column=START_COLUMN_FOR_WRITE_MEASUREMENTS+1, row=START_ROW_FOR_WRITE_MEASUREMENTS+i).value = measurement.amperage
            self._work_page.cell(column=START_COLUMN_FOR_WRITE_MEASUREMENTS+2, row=START_ROW_FOR_WRITE_MEASUREMENTS+i).value = measurement.gas_consumption
            self._work_page.cell(column=END_COLUMN_FOR_WRITE_MEASUREMENTS, row=START_ROW_FOR_WRITE_MEASUREMENTS+i).value = measurement.wire_consumption
        
    
        
        from typing import List, Tuple
        def write_general_parameters(*,
                list_: List[Tuple[str, str]],
                name: str,
                command: str,
                cell_column, 
                cell_row, indent = 0):
            
            for i, (title, ms) in enumerate(list_):
                title = f"{name} {title}, {ms}"
                self._work_page.cell(column=cell_column, row=cell_row+i).value = title
                start_coordinate = self._work_page.cell(column=START_COLUMN_FOR_WRITE_MEASUREMENTS+1+indent+i, row=START_ROW_FOR_WRITE_MEASUREMENTS).coordinate
                end_coordinate = self._work_page.cell(column=START_COLUMN_FOR_WRITE_MEASUREMENTS+1+indent+i, row=END_ROW_FOR_WRITE_MEASUREMENTS).coordinate
                self._work_page.cell(column=cell_column+1, row=cell_row+i).value = f'={command}({start_coordinate}:{end_coordinate})'
        
        
        write_general_parameters(
            list_=titles_and_ms[1:],
            name='Средний',
            command='AVERAGE',
            cell_column=2,
            cell_row=3)
        
        write_general_parameters(
            list_=titles_and_ms[1:],
            name='Максимальный',
            command='MAX',
            cell_column=2,
            cell_row=6)
        
        write_general_parameters(
            list_=titles_and_ms[2:],
            name='Общий',
            command='SUM',
            cell_column=2,
            cell_row=9, indent=1)
        
        
        
        # # calculate AVERAGE values
        # for i, (title, ms) in enumerate(titles_and_ms):
        #     title = "Средний " + title + ', ' + ms
        #     self._work_page.cell(column=2, row=3+i).value = title
        #     start_coordinate = self._work_page.cell(column=START_COLUMN_FOR_WRITE_MEASUREMENTS+1+i, row=START_ROW_FOR_WRITE_MEASUREMENTS).coordinate
        #     end_coordinate = self._work_page.cell(column=START_COLUMN_FOR_WRITE_MEASUREMENTS+1+i, row=END_ROW_FOR_WRITE_MEASUREMENTS).coordinate
        #     self._work_page.cell(column=3, row=3+i).value = f'=AVERAGE({start_coordinate}:{end_coordinate})'
        
        # self.data_store.report.
        
        wpp = self.data_store.welding_process_parameters
        general_parameters = [
            ("Тип металла", wpp.metal_type),
            ("Тип газа", wpp.gas_type),
            ("Плотность металла, кг/м^3", wpp.metal_density),
            ("Диаметр проволки, мм", wpp.wire_diameter),
            
        ]
        for i, (title, value) in enumerate(general_parameters):
            self._work_page.cell(column=4, row=3+i).value = title
            self._work_page.cell(column=5, row=3+i).value = value
            
            
        
        x_time = Reference(
            self._work_page,
            min_col=START_COLUMN_FOR_WRITE_MEASUREMENTS, 
            min_row=START_ROW_FOR_WRITE_MEASUREMENTS, max_row=END_ROW_FOR_WRITE_MEASUREMENTS)
        
        for i, (title, ms) in enumerate(titles_and_ms[1:]):
            print(i)
            y = Reference(
                self._work_page, 
                min_col=START_COLUMN_FOR_WRITE_MEASUREMENTS+1+i, 
                min_row=START_ROW_FOR_WRITE_MEASUREMENTS, max_row=END_ROW_FOR_WRITE_MEASUREMENTS)
            
            s = Series(y, xvalues=x_time)
            print(y)
            chart = ScatterChart()
            chart.title = title
            chart.x_axis.title = ms
            chart.y_axis.title = 'Время'
            chart.legend = None
            chart.append(s)
            self._work_page.add_chart(chart, f"G{(i+1)*20}")
        
        
    def write_measurements(self):
        if not self._did_data_load:
            raise RuntimeError(
                "Данные не были загружены из базы данных"
            )

        
        self._work_page.append(self._option_for_measurement["titles"])

        measurements = self.__measurements
        
        for measurement in measurements:
            self._work_page.append(
                [
                    measurement.utc_time,
                    measurement.amperage,
                    measurement.gas_consumption,
                    measurement.wire_consumption
                ]
            )
        
        x_time = Reference(
            self._work_page,
            min_col=1, 
            min_row=2, max_row=len(measurements))
        
        for i in range(2, len(self._option_for_measurement["titles"])+1):
            
            y = Reference(
                self._work_page, 
                min_col=i, 
                min_row=2, max_row=len(measurements))

            
            s = Series(y, xvalues=x_time)
            
            chart = ScatterChart()
            chart.title = self._option_for_measurement["titles"][i-1]
            chart.x_axis.title = self._option_for_measurement["ms"][0]
            chart.y_axis.title = self._option_for_measurement["ms"][i-1] 
            chart.legend = None
            chart.append(s)
            self._work_page.add_chart(chart, f"G{i*15}")
        
        

        
        for attr in dir(self.__daily_report):
            attr = str(attr)
            if attr.startswith('_') or attr in ["metadata", "sensor"] or attr.endswith("id"):
                continue
            self._option_for_daily_report[attr]["value"] = getattr(self.__daily_report, attr)
        
           
        worker = session.query(Worker).filter(Worker.id==self.__daily_report.worker_id).first()
        self._option_for_daily_report["worker"]["value"] = f"{worker.first_name} {worker.second_name}" \
            if worker else "Пеневайз"
        
        welding_wire_diameter = session.query(WeldingWireDiameter).filter(WeldingWireDiameter.id==self.__daily_report.welding_wire_diameter_id).first()
        self._option_for_daily_report["welding_wire_diameter"]["value"] = welding_wire_diameter.diameter \
            if welding_wire_diameter else "Неизвестно"
        
        weld_metal = session.query(WeldMetal).filter(WeldMetal.id==self.__daily_report.weld_metal_id).first()
        self._option_for_daily_report["weld_metal"]["value"] = weld_metal.steel_name \
            if weld_metal else "Неизвестно"
        
        start_column = 7
        start_row = 1
        for i in self._option_for_daily_report:
            for key in self._option_for_daily_report[i].keys():  
                self._work_page.cell(
                    row=start_row,
                    column=start_column,
                    value=self._option_for_daily_report[i][key]
                )
                start_column += 1
            start_column -= len(self._option_for_daily_report[i].keys())
            start_row += 1
        
    def save(self, path: Path):
        self._work_book.save(self.name_file)
        